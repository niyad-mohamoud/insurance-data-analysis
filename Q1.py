import openpyxl as xl
import numpy as np
import matplotlib.pyplot as plt

wb = xl.load_workbook("customers.xlsx")
sh = wb["Sheet"]

men_count = 0
women_count = 0
claim_men_count = 0
claim_women_count= 0

# Loop through the rows in sh
for row in sh.iter_rows(min_row=2, min_col=4, max_col=8, values_only=True):
    if row is not None:
        if row[0].upper() == "W":
            women_count += 1
            if row[4].upper() == "Y":                
                claim_women_count += 1
        elif row[0].upper() == "M":
            men_count += 1
            if row[4].upper() == "Y":
                claim_men_count += 1



print("Claim probability for men is:", "{:.2f}".format((claim_men_count / men_count * 100)), "%")
print("Claim probability for women is:", "{:.2f}".format((claim_women_count / women_count * 100)), "%")


age_count = np.zeros(20, dtype=int)
claim_count = np.zeros(20, dtype=int)
chance = np.zeros(20, dtype=float)


age_count2 = np.zeros(10, dtype=int)
claim_count2 = np.zeros(10, dtype=int)
chance2 = np.zeros(10, dtype=float)

# Loop through the rows in sh
for row in sh.iter_rows(min_row=2, min_col=4, max_col=8, values_only=True):
    if row is not None:
        age = row[2]
        claim = row[4]
        age_count[int(age / 5)] += 1
        if claim.upper() == "Y":
            claim_count[int(age / 5)] += 1
        
# Calculate the chance of making a claim for each age group
for i in range(len(age_count)):
    if age_count[i] != 0:
        chance[i] = claim_count[i] / age_count[i] * 100
    else:
        chance[i] = 0


for row in sh.iter_rows(min_row=2, min_col=4, max_col=8, values_only=True):
    if row is not None:
        age = row[2]
        claim = row[4]
        age_count2[int(age / 10)] += 1
        if claim.upper() == "Y":
            claim_count2[int(age / 10)] += 1



for i in range(len(age_count2)):
    if age_count2[i] != 0:
        chance2[i] = claim_count2[i] / age_count2[i] * 100
    else:
        chance2[i] = 0

# 5 year intervals
num_groups = 20
range_size = 5

groups = [(i, i + range_size) for i in range(0, num_groups * range_size, range_size)]

plt.style.use("ggplot")

# Plotting the bar chart
fig, ax = plt.subplots()

# Set the x-axis labels to the middle of each group range
x_labels = [f"{g[0]}-{g[1]-1}" for g in groups]

# Set the positions and width for the bars
positions = np.arange(len(groups))

# Plot the bars
ax.bar(positions, chance, align="center", alpha=0.7)

# Set the x-axis labels
ax.set_xticks(positions)
ax.set_xticklabels(x_labels)

# Set the labels and title
ax.set_xlabel("Claim Value Ranges")
ax.set_ylabel("Percentage of Claims (%)")
ax.set_title("Claim Chance by age group with 5 year intervals")

# Show the plot
plt.xticks(rotation=45)  # Rotate x-axis labels if needed
plt.show()



#10 year intervals
num_groups2 = 10
range_size2 = 10

groups2 = [(i, i + range_size2) for i in range(0, num_groups2 * range_size2, range_size2)]


fig2, ax2 = plt.subplots()
# Set the x-axis labels to the middle of each group range
x_labels2 = [f"{g[0]}-{g[1]-1}" for g in groups2]

# Set the positions and width for the bars
positions2 = np.arange(len(groups2))

# Plot the bars
ax2.bar(positions2, chance2, align="center", alpha=0.7)

# Set the x-axis labels
ax2.set_xticks(positions2)
ax2.set_xticklabels(x_labels2)

# Set the labels and title
ax2.set_xlabel("Claim Value Ranges")
ax2.set_ylabel("Percentage of Claims (%)")
ax2.set_title("Claim Chance by age group with 10 year intervals")

# Show the plot
plt.xticks(rotation=45)  # Rotate x-axis labels if needed
plt.show()





